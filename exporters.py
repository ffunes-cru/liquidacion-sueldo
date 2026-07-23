"""
exporters.py — Módulo de importación/exportación y generación de informes (Excel, CSV, PDF, ODS)
para la aplicación de Liquidación de Sueldos.
Soporta importación/exportación plana de variables de empleados (prefijo j_) y variables globales de sistema.
"""

import csv
import json
import os
import sqlite3
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage

from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableRow, TableCell
from odf.text import P
from odf.draw import Frame, Image

from PyQt6.QtGui import QTextDocument, QPageLayout
from PyQt6.QtPrintSupport import QPrinter
from PyQt6.QtCore import QMarginsF


# ======================================================================
# Formateadores para visualización e informes
# ======================================================================
def _formato_moneda(valor) -> str:
    if valor is None:
        return ""
    try:
        v = float(valor)
        signo = "-" if v < 0 else ""
        v = abs(v)
        entero = int(v)
        decimal = round(v - entero, 2)
        entero_str = f"{entero:,}".replace(",", ".")
        decimal_str = f"{decimal:.2f}"[2:]
        return f"{signo}$ {entero_str},{decimal_str}"
    except (ValueError, TypeError):
        return str(valor)


def _formato_porcentaje(valor) -> str:
    if valor is None:
        return ""
    try:
        v = float(valor)
        if abs(v) < 1:
            return f"{v * 100:.2f}%"
        return f"{v:.2f}"
    except (ValueError, TypeError):
        return str(valor)


# ======================================================================
# Gestión de Base de Datos: Excel (Importar/Exportar)
# ======================================================================
def exportar_datos_xlsx(db, path: str):
    """Exporta las tablas de la base de datos a un archivo Excel (.xlsx)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja default

    # 1. Esquemas de Cálculo
    ws_esq = wb.create_sheet(title="Esquemas de Cálculo")
    ws_esq.append(["codigo", "nombre"])
    for row in db.conn.execute("SELECT codigo, nombre FROM esquemas_calculo").fetchall():
        ws_esq.append(list(row))

    # 2. Categorías Jornaleras
    ws_cat = wb.create_sheet(title="Categorías Jornaleras")
    ws_cat.append(["id", "nombre", "valor_hora"])
    for row in db.conn.execute("SELECT id, nombre, valor_hora FROM categorias_jornal").fetchall():
        ws_cat.append(list(row))

    # 3. Secciones
    ws_sec = wb.create_sheet(title="Secciones")
    ws_sec.append(["id", "codigo", "titulo"])
    for row in db.conn.execute("SELECT id, codigo, titulo FROM secciones").fetchall():
        ws_sec.append(list(row))

    # 4. Empleados (Con columnas dinámicas j_ para las variables de cálculo)
    ws_emp = wb.create_sheet(title="Empleados")
    
    # Descubrir todas las claves únicas en los variables_calculo JSON de empleados (incluyendo quincenas de jornaleros)
    todas_claves = set()
    rows_empleados = db.conn.execute("SELECT id, legajo, nombre_completo, tipo_liquidacion, esquema_codigo, categoria_jornal_id, variables_calculo, fecha_ingreso, cuil FROM empleados").fetchall()
    for row in rows_empleados:
        try:
            d = json.loads(row["variables_calculo"])
            if isinstance(d, dict):
                if "quincenas" in d and isinstance(d["quincenas"], dict):
                    for q_vars in d["quincenas"].values():
                        if isinstance(q_vars, dict):
                            todas_claves.update(q_vars.keys())
                else:
                    todas_claves.update(d.keys())
        except Exception:
            pass
    claves_ordenadas = sorted(list(todas_claves))

    # Cabecera
    headers = ["id", "legajo", "nombre_completo", "tipo_liquidacion", "esquema_codigo", "categoria_jornal_id", "fecha_ingreso", "cuil"]
    headers += [f"j_{k}" for k in claves_ordenadas]
    ws_emp.append(headers)

    # Rellenar datos de empleados
    for row in rows_empleados:
        try:
            d_vars = json.loads(row["variables_calculo"])
        except Exception:
            d_vars = {}
            
        es_jornal = row["tipo_liquidacion"] == "jornal"
        
        if es_jornal and isinstance(d_vars, dict) and "quincenas" in d_vars and isinstance(d_vars["quincenas"], dict):
            qs = sorted(list(d_vars["quincenas"].keys()))
            if not qs:
                qs = ["Q1"]
                d_vars["quincenas"] = {"Q1": {}}
            
            for idx, q_code in enumerate(qs):
                q_vars = d_vars["quincenas"].get(q_code, {})
                if idx == 0:
                    vals = [
                        row["id"],
                        row["legajo"],
                        row["nombre_completo"],
                        row["tipo_liquidacion"],
                        row["esquema_codigo"],
                        row["categoria_jornal_id"],
                        row["fecha_ingreso"],
                        row["cuil"]
                    ]
                else:
                    vals = ["", "", "", "", "", "", "", ""]
                
                for k in claves_ordenadas:
                    vals.append(q_vars.get(k, None))
                ws_emp.append(vals)
        else:
            vals = [
                row["id"],
                row["legajo"],
                row["nombre_completo"],
                row["tipo_liquidacion"],
                row["esquema_codigo"],
                row["categoria_jornal_id"],
                row["fecha_ingreso"],
                row["cuil"]
            ]
            for k in claves_ordenadas:
                vals.append(d_vars.get(k, None))
            ws_emp.append(vals)

    # 5. Celdas de Cálculo (Con columnas simples)
    ws_cel = wb.create_sheet(title="Celdas de Cálculo")
    ws_cel.append([
        "id", "seccion_codigo", "codigo_variable", "descripcion", "condicion",
        "formula_unidad", "formula_base", "formula_monto", "orden", "esquema_codigo",
        "tipo_calculo", "simple_porcentaje", "simple_base_variable", "simple_monto_fijo"
    ])
    for row in db.conn.execute(
        """SELECT id, seccion_codigo, codigo_variable, descripcion, condicion,
                  formula_unidad, formula_base, formula_monto, orden, esquema_codigo,
                  tipo_calculo, simple_porcentaje, simple_base_variable, simple_monto_fijo
           FROM celdas_calculo"""
    ).fetchall():
        ws_cel.append(list(row))

    # 6. Celdas de Gráfico Custom
    ws_g = wb.create_sheet(title="Celdas de Gráfico")
    ws_g.append(["id", "etiqueta", "formula", "orden", "esquema_codigo"])
    for row in db.conn.execute("SELECT id, etiqueta, formula, orden, esquema_codigo FROM celdas_grafico").fetchall():
        ws_g.append(list(row))

    # 7. Variables Globales
    ws_glob = wb.create_sheet(title="Variables Globales")
    ws_glob.append(["id", "codigo", "valor", "descripcion"])
    for row in db.conn.execute("SELECT id, codigo, valor, descripcion FROM variables_globales").fetchall():
        ws_glob.append(list(row))

    # 8. Empresa
    ws_empr = wb.create_sheet(title="Empresa")
    ws_empr.append(["id", "razon_social", "direccion", "cuit", "lugar_de_pago"])
    for row in db.conn.execute("SELECT id, razon_social, direccion, cuit, lugar_de_pago FROM empresa").fetchall():
        ws_empr.append(list(row))

    # 9. Recibos
    ws_rec = wb.create_sheet(title="Recibos")
    ws_rec.append(["id", "empleado_id", "esquema_codigo", "mes", "anio", "periodo", "datos_json", "fecha_emision"])
    for row in db.conn.execute("SELECT id, empleado_id, esquema_codigo, mes, anio, periodo, datos_json, fecha_emision FROM recibos").fetchall():
        ws_rec.append(list(row))

    wb.save(path)


def importar_datos_xlsx(db, path: str):
    """Importa datos desde un archivo Excel (.xlsx), reemplazando las tablas actuales."""
    wb = openpyxl.load_workbook(path)
    cur = db.conn.cursor()

    try:
        cur.execute("PRAGMA foreign_keys = OFF")
        cur.execute("BEGIN TRANSACTION")

        # Importar Esquemas
        if "Esquemas de Cálculo" in wb.sheetnames:
            cur.execute("DELETE FROM celdas_calculo")  # Limpiar dependencias
            cur.execute("DELETE FROM celdas_grafico")
            cur.execute("DELETE FROM empleados")
            cur.execute("DELETE FROM esquemas_calculo")
            sheet = wb["Esquemas de Cálculo"]
            rows = list(sheet.rows)[1:]
            for row in rows:
                vals = [cell.value for cell in row]
                if vals[0] is not None:
                    cur.execute("INSERT INTO esquemas_calculo (codigo, nombre) VALUES (?, ?)", (vals[0], vals[1]))

        # Importar Categorías
        if "Categorías Jornaleras" in wb.sheetnames:
            cur.execute("DELETE FROM categorias_jornal")
            sheet = wb["Categorías Jornaleras"]
            rows = list(sheet.rows)[1:]
            for row in rows:
                vals = [cell.value for cell in row]
                if vals[1] is not None:
                    cur.execute("INSERT INTO categorias_jornal (id, nombre, valor_hora) VALUES (?, ?, ?)", (vals[0], vals[1], vals[2]))

        # Importar Secciones
        if "Secciones" in wb.sheetnames:
            cur.execute("DELETE FROM secciones")
            sheet = wb["Secciones"]
            rows = list(sheet.rows)[1:]
            for row in rows:
                vals = [cell.value for cell in row]
                if vals[1] is not None:
                    cur.execute("INSERT INTO secciones (id, codigo, titulo) VALUES (?, ?, ?)", (vals[0], vals[1], vals[2]))

        # Importar Empleados con mapeo de columnas dinámicas j_
        if "Empleados" in wb.sheetnames:
            cur.execute("DELETE FROM empleados")
            sheet = wb["Empleados"]
            rows = list(sheet.rows)
            if len(rows) > 0:
                headers = [cell.value for cell in rows[0]]
                id_idx = headers.index("id")
                legajo_idx = headers.index("legajo")
                nombre_idx = headers.index("nombre_completo")
                tipo_idx = headers.index("tipo_liquidacion")
                esquema_idx = headers.index("esquema_codigo") if "esquema_codigo" in headers else -1
                cat_idx = headers.index("categoria_jornal_id") if "categoria_jornal_id" in headers else -1
                fecha_ingreso_idx = headers.index("fecha_ingreso") if "fecha_ingreso" in headers else -1
                cuil_idx = headers.index("cuil") if "cuil" in headers else -1
                
                j_cols = {}
                for idx, h in enumerate(headers):
                    if h and h.startswith("j_"):
                        j_cols[h[2:]] = idx
 
                current_emp = None
                current_jornal_quincenas = []
 
                def save_current_emp():
                    if current_emp is None:
                        return
                    if current_emp["tipo_liquidacion"] == "jornal":
                        q_dict = {}
                        for idx, q_vars in enumerate(current_jornal_quincenas):
                            q_code = f"Q{idx+1}"
                            q_dict[q_code] = q_vars
                        variables_json = json.dumps({"quincenas": q_dict})
                    else:
                        variables_json = json.dumps(current_emp["variables_calculo"])
                    
                    cur.execute(
                        """INSERT INTO empleados (id, legajo, nombre_completo, tipo_liquidacion, variables_calculo, esquema_codigo, categoria_jornal_id, fecha_ingreso, cuil)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (current_emp["id"], current_emp["legajo"], current_emp["nombre_completo"], current_emp["tipo_liquidacion"],
                         variables_json, current_emp["esquema_codigo"], current_emp["categoria_jornal_id"], current_emp["fecha_ingreso"], current_emp["cuil"])
                    )
 
                for row in rows[1:]:
                    vals = [cell.value for cell in row]
                    if len(vals) <= nombre_idx:
                        continue
                    
                    is_new_emp = vals[nombre_idx] is not None and str(vals[nombre_idx]).strip() != ""
                    
                    d_vars = {}
                    for var_name, col_i in j_cols.items():
                        if col_i < len(vals):
                            val = vals[col_i]
                            if val is not None and val != "":
                                if isinstance(val, str) and (val.startswith("{") or val.startswith("[")):
                                    try:
                                        val = json.loads(val)
                                    except Exception:
                                        pass
                                d_vars[var_name] = val
 
                    if is_new_emp:
                        save_current_emp()
                        current_emp = {
                            "id": vals[id_idx],
                            "legajo": vals[legajo_idx],
                            "nombre_completo": vals[nombre_idx],
                            "tipo_liquidacion": vals[tipo_idx],
                            "esquema_codigo": vals[esquema_idx] if (esquema_idx >= 0 and vals[esquema_idx]) else 'MENSUAL',
                            "categoria_jornal_id": vals[cat_idx] if (cat_idx >= 0 and vals[cat_idx] is not None) else None,
                            "fecha_ingreso": vals[fecha_ingreso_idx] if (fecha_ingreso_idx >= 0 and vals[fecha_ingreso_idx]) else '2020-01-01',
                            "cuil": vals[cuil_idx] if (cuil_idx >= 0 and vals[cuil_idx]) else '',
                            "variables_calculo": d_vars
                        }
                        current_jornal_quincenas = [d_vars]
                    else:
                        if current_emp and current_emp["tipo_liquidacion"] == "jornal":
                            current_jornal_quincenas.append(d_vars)
                save_current_emp()

        # Importar Celdas Cálculo
        if "Celdas de Cálculo" in wb.sheetnames:
            cur.execute("DELETE FROM celdas_calculo")
            sheet = wb["Celdas de Cálculo"]
            rows = list(sheet.rows)[1:]
            for row in rows:
                vals = [cell.value for cell in row]
                if vals[2] is not None:  # codigo_variable
                    cur.execute(
                        """INSERT INTO celdas_calculo
                           (id, seccion_codigo, codigo_variable, descripcion, condicion,
                            formula_unidad, formula_base, formula_monto, orden, esquema_codigo,
                            tipo_calculo, simple_porcentaje, simple_base_variable, simple_monto_fijo)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (vals[0], vals[1], vals[2], vals[3], vals[4] or '',
                         vals[5] or '', vals[6] or '', vals[7] or '', vals[8] or 0,
                         vals[9] or 'MENSUAL', vals[10] or 'formula', vals[11], vals[12], vals[13])
                    )

        # Importar Celdas Gráfico Custom
        if "Celdas de Gráfico" in wb.sheetnames:
            cur.execute("DELETE FROM celdas_grafico")
            sheet = wb["Celdas de Gráfico"]
            rows = list(sheet.rows)[1:]
            for row in rows:
                vals = [cell.value for cell in row]
                if vals[1] is not None:  # etiqueta
                    cur.execute(
                        "INSERT INTO celdas_grafico (id, etiqueta, formula, orden, esquema_codigo) VALUES (?, ?, ?, ?, ?)",
                        (vals[0], vals[1], vals[2], vals[3] or 0, vals[4] or 'MENSUAL')
                    )

        # Importar Variables Globales (NUEVO)
        if "Variables Globales" in wb.sheetnames:
            cur.execute("DELETE FROM variables_globales")
            sheet = wb["Variables Globales"]
            rows = list(sheet.rows)[1:]
            for row in rows:
                vals = [cell.value for cell in row]
                if vals[1] is not None:  # codigo
                    cur.execute(
                        "INSERT INTO variables_globales (id, codigo, valor, descripcion) VALUES (?, ?, ?, ?)",
                        (vals[0], vals[1], str(vals[2]), vals[3] or '')
                    )

        # Importar Empresa
        if "Empresa" in wb.sheetnames:
            cur.execute("DELETE FROM empresa")
            sheet = wb["Empresa"]
            rows = list(sheet.rows)[1:]
            for row in rows:
                vals = [cell.value for cell in row]
                if vals[0] is not None:
                    cur.execute("INSERT INTO empresa (id, razon_social, direccion, cuit, lugar_de_pago) VALUES (?, ?, ?, ?, ?)",
                                (vals[0], vals[1] or '', vals[2] or '', vals[3] or '', vals[4] or ''))

        # Importar Recibos
        if "Recibos" in wb.sheetnames:
            cur.execute("DELETE FROM recibos")
            sheet = wb["Recibos"]
            rows = list(sheet.rows)[1:]
            for row in rows:
                vals = [cell.value for cell in row]
                if vals[0] is not None:
                    cur.execute("INSERT INTO recibos (id, empleado_id, esquema_codigo, mes, anio, periodo, datos_json, fecha_emision) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                                (vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], vals[6] or '{}', vals[7] or ''))

        db.conn.commit()
    except Exception as e:
        db.conn.rollback()
        raise e
    finally:
        cur.execute("PRAGMA foreign_keys = ON")


# ======================================================================
# Gestión de Base de Datos: CSV (Importar/Exportar)
# ======================================================================
def exportar_datos_csv(db, directorio: str):
    """Exporta todas las tablas a archivos CSV en el directorio indicado."""
    os.makedirs(directorio, exist_ok=True)

    # 1. Esquemas
    with open(os.path.join(directorio, "esquemas_calculo.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["codigo", "nombre"])
        for row in db.conn.execute("SELECT codigo, nombre FROM esquemas_calculo").fetchall():
            writer.writerow(list(row))

    # 2. Categorías
    with open(os.path.join(directorio, "categorias_jornal.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "nombre", "valor_hora"])
        for row in db.conn.execute("SELECT id, nombre, valor_hora FROM categorias_jornal").fetchall():
            writer.writerow(list(row))

    # 3. Secciones
    with open(os.path.join(directorio, "secciones.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "codigo", "titulo"])
        for row in db.conn.execute("SELECT id, codigo, titulo FROM secciones").fetchall():
            writer.writerow(list(row))

    # 4. Empleados (Con columnas dinámicas j_)
    todas_claves = set()
    rows_empleados = db.conn.execute("SELECT id, legajo, nombre_completo, tipo_liquidacion, esquema_codigo, categoria_jornal_id, variables_calculo, fecha_ingreso, cuil FROM empleados").fetchall()
    for row in rows_empleados:
        try:
            d = json.loads(row["variables_calculo"])
            if isinstance(d, dict):
                if "quincenas" in d and isinstance(d["quincenas"], dict):
                    for q_vars in d["quincenas"].values():
                        if isinstance(q_vars, dict):
                            todas_claves.update(q_vars.keys())
                else:
                    todas_claves.update(d.keys())
        except Exception:
            pass
    claves_ordenadas = sorted(list(todas_claves))

    with open(os.path.join(directorio, "empleados.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        headers = ["id", "legajo", "nombre_completo", "tipo_liquidacion", "esquema_codigo", "categoria_jornal_id", "fecha_ingreso", "cuil"]
        headers += [f"j_{k}" for k in claves_ordenadas]
        writer.writerow(headers)

        for row in rows_empleados:
            try:
                d_vars = json.loads(row["variables_calculo"])
            except Exception:
                d_vars = {}
                
            es_jornal = row["tipo_liquidacion"] == "jornal"
            
            if es_jornal and isinstance(d_vars, dict) and "quincenas" in d_vars and isinstance(d_vars["quincenas"], dict):
                qs = sorted(list(d_vars["quincenas"].keys()))
                if not qs:
                    qs = ["Q1"]
                    d_vars["quincenas"] = {"Q1": {}}
                
                for idx, q_code in enumerate(qs):
                    q_vars = d_vars["quincenas"].get(q_code, {})
                    if idx == 0:
                        vals = [
                            row["id"],
                            row["legajo"],
                            row["nombre_completo"],
                            row["tipo_liquidacion"],
                            row["esquema_codigo"],
                            row["categoria_jornal_id"],
                            row["fecha_ingreso"],
                            row["cuil"]
                        ]
                    else:
                        vals = ["", "", "", "", "", "", "", ""]
                    
                    for k in claves_ordenadas:
                        vals.append(q_vars.get(k, ""))
                    writer.writerow(vals)
            else:
                vals = [
                    row["id"],
                    row["legajo"],
                    row["nombre_completo"],
                    row["tipo_liquidacion"],
                    row["esquema_codigo"],
                    row["categoria_jornal_id"],
                    row["fecha_ingreso"],
                    row["cuil"]
                ]
                for k in claves_ordenadas:
                    vals.append(d_vars.get(k, ""))
                writer.writerow(vals)

    # 5. Celdas Cálculo
    with open(os.path.join(directorio, "celdas_calculo.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "id", "seccion_codigo", "codigo_variable", "descripcion", "condicion",
            "formula_unidad", "formula_base", "formula_monto", "orden", "esquema_codigo",
            "tipo_calculo", "simple_porcentaje", "simple_base_variable", "simple_monto_fijo"
        ])
        for row in db.conn.execute(
            """SELECT id, seccion_codigo, codigo_variable, descripcion, condicion,
                      formula_unidad, formula_base, formula_monto, orden, esquema_codigo,
                      tipo_calculo, simple_porcentaje, simple_base_variable, simple_monto_fijo
               FROM celdas_calculo"""
        ).fetchall():
            writer.writerow(list(row))

    # 6. Celdas Gráfico
    with open(os.path.join(directorio, "celdas_grafico.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "etiqueta", "formula", "orden", "esquema_codigo"])
        for row in db.conn.execute("SELECT id, etiqueta, formula, orden, esquema_codigo FROM celdas_grafico").fetchall():
            writer.writerow(list(row))

    # 7. Variables Globales (NUEVO)
    with open(os.path.join(directorio, "variables_globales.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "codigo", "valor", "descripcion"])
        for row in db.conn.execute("SELECT id, codigo, valor, descripcion FROM variables_globales").fetchall():
            writer.writerow(list(row))

    # 8. Empresa
    with open(os.path.join(directorio, "empresa.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "razon_social", "direccion", "cuit", "lugar_de_pago"])
        for row in db.conn.execute("SELECT id, razon_social, direccion, cuit, lugar_de_pago FROM empresa").fetchall():
            writer.writerow(list(row))

    # 9. Recibos
    with open(os.path.join(directorio, "recibos.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "empleado_id", "esquema_codigo", "mes", "anio", "periodo", "datos_json", "fecha_emision"])
        for row in db.conn.execute("SELECT id, empleado_id, esquema_codigo, mes, anio, periodo, datos_json, fecha_emision FROM recibos").fetchall():
            writer.writerow(list(row))


def importar_datos_csv(db, directorio: str):
    """Importa datos desde los CSVs en el directorio, reemplazando datos actuales."""
    cur = db.conn.cursor()
    try:
        cur.execute("PRAGMA foreign_keys = OFF")
        cur.execute("BEGIN TRANSACTION")

        # 1. Esquemas
        p_esq = os.path.join(directorio, "esquemas_calculo.csv")
        if os.path.exists(p_esq):
            cur.execute("DELETE FROM celdas_calculo")
            cur.execute("DELETE FROM celdas_grafico")
            cur.execute("DELETE FROM empleados")
            cur.execute("DELETE FROM esquemas_calculo")
            with open(p_esq, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for vals in reader:
                    if vals:
                        cur.execute("INSERT INTO esquemas_calculo (codigo, nombre) VALUES (?, ?)", (vals[0], vals[1]))

        # 2. Categorías
        p_cat = os.path.join(directorio, "categorias_jornal.csv")
        if os.path.exists(p_cat):
            cur.execute("DELETE FROM categorias_jornal")
            with open(p_cat, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for vals in reader:
                    if vals:
                        cur.execute("INSERT INTO categorias_jornal (id, nombre, valor_hora) VALUES (?, ?, ?)", (vals[0], vals[1], vals[2]))

        # 3. Secciones
        p_sec = os.path.join(directorio, "secciones.csv")
        if os.path.exists(p_sec):
            cur.execute("DELETE FROM secciones")
            with open(p_sec, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for vals in reader:
                    if vals:
                        cur.execute("INSERT INTO secciones (id, codigo, titulo) VALUES (?, ?, ?)", (vals[0], vals[1], vals[2]))

        # 4. Empleados con variables dinámicas j_
        p_emp = os.path.join(directorio, "empleados.csv")
        if os.path.exists(p_emp):
            cur.execute("DELETE FROM empleados")
            with open(p_emp, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                
                current_emp = None
                current_jornal_quincenas = []
 
                def save_current_emp():
                    if current_emp is None:
                        return
                    if current_emp["tipo_liquidacion"] == "jornal":
                        q_dict = {}
                        for idx, q_vars in enumerate(current_jornal_quincenas):
                            q_code = f"Q{idx+1}"
                            q_dict[q_code] = q_vars
                        variables_json = json.dumps({"quincenas": q_dict})
                    else:
                        variables_json = json.dumps(current_emp["variables_calculo"])
                    
                    cur.execute(
                        """INSERT INTO empleados (id, legajo, nombre_completo, tipo_liquidacion, variables_calculo, esquema_codigo, categoria_jornal_id, fecha_ingreso, cuil)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (current_emp["id"], current_emp["legajo"], current_emp["nombre_completo"], current_emp["tipo_liquidacion"],
                         variables_json, current_emp["esquema_codigo"], current_emp["categoria_jornal_id"], current_emp["fecha_ingreso"], current_emp["cuil"])
                    )
 
                for row in reader:
                    nombre_val = row.get("nombre_completo", "")
                    id_val = row.get("id", "")
                    
                    is_new_emp = nombre_val is not None and str(nombre_val).strip() != "" and id_val is not None and str(id_val).strip() != ""
                    
                    d_vars = {}
                    for k, val in row.items():
                        if k.startswith("j_") and val != "" and val is not None:
                            if isinstance(val, str) and (val.startswith("{") or val.startswith("[")):
                                try:
                                    d_vars[k[2:]] = json.loads(val)
                                    continue
                                except Exception:
                                    pass
                            try:
                                if "." in val:
                                    d_vars[k[2:]] = float(val)
                                else:
                                    d_vars[k[2:]] = int(val)
                            except ValueError:
                                if val.lower() == "true":
                                    d_vars[k[2:]] = True
                                elif val.lower() == "false":
                                    d_vars[k[2:]] = False
                                else:
                                    d_vars[k[2:]] = val
 
                    if is_new_emp:
                        save_current_emp()
                        current_emp = {
                            "id": row["id"],
                            "legajo": row["legajo"],
                            "nombre_completo": row["nombre_completo"],
                            "tipo_liquidacion": row["tipo_liquidacion"],
                            "esquema_codigo": row.get("esquema_codigo", "MENSUAL"),
                            "categoria_jornal_id": row.get("categoria_jornal_id") if row.get("categoria_jornal_id") else None,
                            "fecha_ingreso": row.get("fecha_ingreso") if row.get("fecha_ingreso") else '2020-01-01',
                            "cuil": row.get("cuil") if row.get("cuil") else '',
                            "variables_calculo": d_vars
                        }
                        current_jornal_quincenas = [d_vars]
                    else:
                        if current_emp and current_emp["tipo_liquidacion"] == "jornal":
                            current_jornal_quincenas.append(d_vars)
                save_current_emp()

        # 5. Celdas Cálculo
        p_cel = os.path.join(directorio, "celdas_calculo.csv")
        if os.path.exists(p_cel):
            cur.execute("DELETE FROM celdas_calculo")
            with open(p_cel, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for vals in reader:
                    if vals:
                        cur.execute(
                            """INSERT INTO celdas_calculo
                               (id, seccion_codigo, codigo_variable, descripcion, condicion,
                                formula_unidad, formula_base, formula_monto, orden, esquema_codigo,
                                tipo_calculo, simple_porcentaje, simple_base_variable, simple_monto_fijo)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], vals[6], vals[7], vals[8],
                             vals[9] or 'MENSUAL', vals[10] or 'formula', vals[11] or None, vals[12] or None, vals[13] or None)
                        )

        # 6. Celdas Gráfico
        p_cg = os.path.join(directorio, "celdas_grafico.csv")
        if os.path.exists(p_cg):
            cur.execute("DELETE FROM celdas_grafico")
            with open(p_cg, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for vals in reader:
                    if vals:
                        cur.execute(
                            "INSERT INTO celdas_grafico (id, etiqueta, formula, orden, esquema_codigo) VALUES (?, ?, ?, ?, ?)",
                            (vals[0], vals[1], vals[2], vals[3], vals[4])
                        )

        # 7. Variables Globales (NUEVO)
        p_glob = os.path.join(directorio, "variables_globales.csv")
        if os.path.exists(p_glob):
            cur.execute("DELETE FROM variables_globales")
            with open(p_glob, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for vals in reader:
                    if vals:
                        cur.execute(
                            "INSERT INTO variables_globales (id, codigo, valor, descripcion) VALUES (?, ?, ?, ?)",
                            (vals[0], vals[1], vals[2], vals[3])
                        )

        # 8. Empresa
        p_empresa = os.path.join(directorio, "empresa.csv")
        if os.path.exists(p_empresa):
            cur.execute("DELETE FROM empresa")
            with open(p_empresa, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for vals in reader:
                    if vals:
                        cur.execute(
                            "INSERT INTO empresa (id, razon_social, direccion, cuit, lugar_de_pago) VALUES (?, ?, ?, ?, ?)",
                            (vals[0], vals[1], vals[2], vals[3], vals[4])
                        )

        # 9. Recibos
        p_recibos = os.path.join(directorio, "recibos.csv")
        if os.path.exists(p_recibos):
            cur.execute("DELETE FROM recibos")
            with open(p_recibos, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for vals in reader:
                    if vals:
                        cur.execute(
                            "INSERT INTO recibos (id, empleado_id, esquema_codigo, mes, anio, periodo, datos_json, fecha_emision) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                            (vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], vals[6] or '{}', vals[7])
                        )

        db.conn.commit()
    except Exception as e:
        db.conn.rollback()
        raise e
    finally:
        cur.execute("PRAGMA foreign_keys = ON")


# ======================================================================
# Generación de Gráfico de Torta (Matplotlib)
# ======================================================================
def generar_grafico_torta(resultado: dict, seccion_codigo: str, path_imagen: str) -> bool:
    """Genera un archivo de gráfico de torta PNG para la sección seleccionada o CUSTOM."""
    if seccion_codigo == "CUSTOM":
        items = resultado.get("resultados_grafico_custom", [])
        items = [f for f in items if f["valor"] > 0]
        if not items:
            return False
        labels = [f["etiqueta"] for f in items]
        valores = [f["valor"] for f in items]
    else:
        filas = resultado["resultados_por_seccion"].get(seccion_codigo, [])
        if not filas:
            return False
        ignorar = ("bruto", "total_deducciones", "neto", "total_cargas_patronales", "costo_laboral_total")
        items = [f for f in filas if f["codigo"] not in ignorar and f["monto"] > 0 and f.get("visible_recibo", 1) == 1]
        if not items:
            return False
        labels = [f["descripcion"] for f in items]
        valores = [f["monto"] for f in items]

    fig, ax = plt.subplots(figsize=(6, 4.5), dpi=100)
    
    def make_autopct(values):
        def my_autopct(pct):
            total = sum(values)
            val = pct * total / 100.0
            return f'{pct:.1f}%\n(${val:,.2f})'
        return my_autopct

    wedges, texts, autotexts = ax.pie(
        valores,
        labels=labels,
        autopct=make_autopct(valores),
        startangle=140,
        textprops=dict(color="black", size=9)
    )
    
    for autotext in autotexts:
        autotext.set_fontsize(8)

    ax.axis("equal")
    
    titulos = {
        "COMPOSICION": "Distribución de Ingresos (Composición Salarial)",
        "RECIBO": "Distribución de Retenciones / Deducciones",
        "COSTO_EMP": "Distribución de Cargas Patronales",
        "CUSTOM": "Distribución de Costos Personalizada"
    }
    ax.set_title(titulos.get(seccion_codigo, "Distribución de Costos"), fontsize=11, fontweight='bold', pad=15)
    
    plt.tight_layout()
    fig.savefig(path_imagen, bbox_inches="tight", transparent=False)
    plt.close(fig)
    return True


# ======================================================================
# Exportar Recibo: PDF (Usa QPrinter/QTextDocument con imagen embebida)
# ======================================================================
def exportar_recibo_pdf(resultado: dict, db, pdf_path: str, path_grafico: str | None,
                        empresa: dict | None = None, mes_anio: dict | None = None):
    """Exporta el recibo detallado a un documento PDF con cabecera industrial."""
    emp = resultado["empleado"]
    secciones_info = {s["codigo"]: s["titulo"] for s in db.listar_secciones()}

    if empresa is None:
        empresa = db.obtener_empresa()
    if mes_anio is None:
        from datetime import date
        hoy = date.today()
        mes_anio = {"mes": hoy.month, "anio": hoy.year, "periodo": "M"}

    esquema = emp.get("esquema_codigo") or "MENSUAL"
    orden_secciones = [s["codigo"] for s in db.listar_secciones()]
    for s_cod in resultado["resultados_por_seccion"].keys():
        if s_cod not in orden_secciones:
            orden_secciones.append(s_cod)

    # Construir filas de la tabla en HTML
    rows_html = []
    for sec_codigo in orden_secciones:
        filas = resultado["resultados_por_seccion"].get(sec_codigo, [])
        filas_visibles = []
        for f in filas:
            if f.get("visible_recibo", 1) == 1:
                es_total = f["codigo"].startswith("total_") or f["codigo"] in ("bruto", "total_deducciones", "neto", "total_cargas_patronales", "costo_laboral_total")
                if f["monto"] != 0 or es_total:
                    filas_visibles.append(f)
        if not filas_visibles:
            continue

        rows_html.append(f"""
            <tr style="background-color: #E2E8F0;">
                <td colspan="4" style="font-weight: bold; color: #1E293B; font-size: 11px;">{secciones_info.get(sec_codigo, sec_codigo).upper()}</td>
            </tr>
        """)

        for f in filas_visibles:
            es_total = f["codigo"].startswith("total_") or f["codigo"] in ("bruto", "total_deducciones", "neto", "total_cargas_patronales", "costo_laboral_total")
            style_row = "font-weight: bold; background-color: #F8FAFC;" if es_total else ""

            u_val = _formato_porcentaje(f["unidad"]) if f["unidad"] is not None else ""
            b_val = _formato_moneda(f["base"]) if f["base"] is not None else ""
            m_val = _formato_moneda(f["monto"])

            rows_html.append(f"""
                <tr style="{style_row}">
                    <td style="padding: 5px; border-bottom: 1px solid #CBD5E1;">{f["descripcion"]}</td>
                    <td style="padding: 5px; border-bottom: 1px solid #CBD5E1; text-align: right;">{u_val}</td>
                    <td style="padding: 5px; border-bottom: 1px solid #CBD5E1; text-align: right;">{b_val}</td>
                    <td style="padding: 5px; border-bottom: 1px solid #CBD5E1; text-align: right;">{m_val}</td>
                </tr>
            """)

    tabla_rows = "\n".join(rows_html)

    img_html = ""
    if path_grafico and os.path.exists(path_grafico):
        img_html = f"""
        <div style="text-align: center; margin-top: 20px; page-break-inside: avoid;">
            <img src="{path_grafico}" width="450" />
        </div>
        """

    # Datos de cabecera
    razon_social = empresa.get("razon_social", "") or "—"
    direccion = empresa.get("direccion", "") or ""
    cuit_emp = empresa.get("cuit", "") or ""
    lugar_pago = empresa.get("lugar_de_pago", "") or ""
    cuil = emp.get("cuil", "") or ""
    fecha_ingreso = emp.get("fecha_ingreso", "") or ""
    cat_nombre = emp.get("categoria_nombre", "") or ""

    # Antigüedad
    from motor import calcular_antiguedad_anios
    from datetime import date
    antiguedad = calcular_antiguedad_anios(fecha_ingreso, date.today().strftime("%Y-%m-%d")) if fecha_ingreso else 0

    # Período
    meses_nombres = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    mes_num = mes_anio.get("mes", 1)
    mes_nombre = meses_nombres[mes_num] if 1 <= mes_num <= 12 else str(mes_num)
    anio_str = str(mes_anio.get("anio", 2025))
    periodo = mes_anio.get("periodo", "M")
    periodo_str = f"Quincena {periodo}" if (periodo and periodo != "M") else "Mensual"

    html_content = f"""
    <html>
    <head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            color: #334155;
            margin: 15px;
            font-size: 11px;
        }}
        .header-empresa {{
            border: 2px solid #1E293B;
            padding: 10px 15px;
            margin-bottom: 12px;
        }}
        .header-empresa h2 {{
            margin: 0 0 4px 0;
            font-size: 16px;
            color: #0F172A;
        }}
        .header-empresa .subtitle {{
            color: #64748B;
            font-size: 10px;
            margin: 0;
        }}
        .info-grid {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 12px;
            font-size: 11px;
        }}
        .info-grid td {{
            padding: 3px 8px;
            border: 1px solid #E2E8F0;
        }}
        .info-grid .label {{
            font-weight: bold;
            background-color: #F8FAFC;
            width: 20%;
            color: #475569;
        }}
        .periodo-bar {{
            background-color: #1E293B;
            color: #FFFFFF;
            padding: 6px 12px;
            font-weight: bold;
            font-size: 12px;
            margin-bottom: 12px;
            text-align: center;
        }}
        .receipt-table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 15px;
            font-size: 11px;
        }}
        .receipt-table th {{
            background-color: #475569;
            color: #FFFFFF;
            font-weight: bold;
            text-align: left;
            padding: 5px 6px;
            font-size: 10px;
        }}
    </style>
    </head>
    <body>
        <!-- Cabecera Empresa -->
        <div class="header-empresa">
            <h2>RECIBO DE HABERES</h2>
            <table style="width:100%; border-collapse:collapse; font-size:11px;">
                <tr>
                    <td style="width:60%;">
                        <b>{razon_social}</b><br/>
                        {f'{direccion}<br/>' if direccion else ''}
                        {f'CUIT: {cuit_emp}' if cuit_emp else ''}
                    </td>
                    <td style="width:40%; text-align:right; vertical-align:top;">
                        {f'Lugar de Pago: {lugar_pago}<br/>' if lugar_pago else ''}
                        Esquema: {esquema}
                    </td>
                </tr>
            </table>
        </div>

        <!-- Barra de Período -->
        <div class="periodo-bar">
            {mes_nombre} {anio_str} — {periodo_str}
        </div>

        <!-- Datos del Empleado -->
        <table class="info-grid">
            <tr>
                <td class="label">Legajo</td>
                <td>{emp["legajo"]}</td>
                <td class="label">CUIL</td>
                <td>{cuil}</td>
            </tr>
            <tr>
                <td class="label">Apellido y Nombre</td>
                <td>{emp["nombre_completo"]}</td>
                <td class="label">Categoría</td>
                <td>{cat_nombre}</td>
            </tr>
            <tr>
                <td class="label">Fecha de Ingreso</td>
                <td>{fecha_ingreso}</td>
                <td class="label">Antigüedad</td>
                <td>{antiguedad} año(s)</td>
            </tr>
            <tr>
                <td class="label">Tipo Liquidación</td>
                <td>{emp["tipo_liquidacion"].capitalize()}</td>
                <td class="label"></td>
                <td></td>
            </tr>
        </table>

        <!-- Tabla del Recibo -->
        <table class="receipt-table">
            <thead>
                <tr>
                    <th width="45%">Concepto</th>
                    <th width="15%" style="text-align: right;">Unidad / %</th>
                    <th width="20%" style="text-align: right;">Base</th>
                    <th width="20%" style="text-align: right;">Monto</th>
                </tr>
            </thead>
            <tbody>
                {tabla_rows}
            </tbody>
        </table>

        {img_html}
    </body>
    </html>
    """

    doc = QTextDocument()
    doc.setHtml(html_content)

    printer = QPrinter(QPrinter.PrinterMode.ScreenResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(pdf_path)
    printer.setPageMargins(QMarginsF(12, 12, 12, 12), QPageLayout.Unit.Millimeter)

    doc.print(printer)


# ======================================================================
# Exportar Recibo: Excel (openpyxl)
# ======================================================================
def exportar_recibo_xlsx(resultado: dict, db, path: str, path_grafico: str | None):
    """Exporta el recibo a formato Excel (.xlsx) con estilos y gráfico."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recibo de Sueldo"

    font_title = Font(name="Calibri", size=14, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_normal = Font(name="Calibri", size=10)
    fill_header = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
    fill_section = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:D1")
    ws["A1"] = "RECIBO DE SUELDO Y COSTO LABORAL"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 25

    emp = resultado["empleado"]
    ws["A3"] = "Legajo:"
    ws["A3"].font = font_bold
    ws["B3"] = emp["legajo"]
    ws["B3"].font = font_normal

    ws["C3"] = "Tipo Liquidación:"
    ws["C3"].font = font_bold
    ws["D3"] = emp["tipo_liquidacion"].capitalize()
    ws["D3"].font = font_normal

    ws["A4"] = "Empleado:"
    ws["A4"].font = font_bold
    ws["B4"] = emp["nombre_completo"]
    ws["B4"].font = font_normal

    if resultado.get("quincena_sel"):
        ws["C4"] = "Quincena:"
        ws["C4"].font = font_bold
        ws["D4"] = resultado["quincena_sel"]
        ws["D4"].font = font_normal

    headers = ["Concepto", "Unidad / %", "Base", "Monto"]
    ws.row_dimensions[6].height = 20
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 1 else align_left

    secciones_info = {s["codigo"]: s["titulo"] for s in db.listar_secciones()}
    orden_secciones = [s["codigo"] for s in db.listar_secciones()]
    for s_cod in resultado["resultados_por_seccion"].keys():
        if s_cod not in orden_secciones:
            orden_secciones.append(s_cod)

    row_num = 7
    for sec_codigo in orden_secciones:
        filas = resultado["resultados_por_seccion"].get(sec_codigo, [])
        filas_visibles = []
        for f in filas:
            if f.get("visible_recibo", 1) == 1:
                es_total = f["codigo"].startswith("total_") or f["codigo"] in ("bruto", "total_deducciones", "neto", "total_cargas_patronales", "costo_laboral_total")
                if f["monto"] != 0 or es_total:
                    filas_visibles.append(f)
        if not filas_visibles:
            continue

        ws.row_dimensions[row_num].height = 20
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        sec_cell = ws.cell(row=row_num, column=1, value=secciones_info.get(sec_codigo, sec_codigo).upper())
        sec_cell.font = font_bold
        for col_i in range(1, 5):
            ws.cell(row=row_num, column=col_i).fill = fill_section
        row_num += 1

        for f in filas_visibles:
            es_total = f["codigo"].startswith("total_") or f["codigo"] in ("bruto", "total_deducciones", "neto", "total_cargas_patronales", "costo_laboral_total")
            
            c_desc = ws.cell(row=row_num, column=1, value=f["descripcion"])
            c_u = ws.cell(row=row_num, column=2, value=f["unidad"])
            c_b = ws.cell(row=row_num, column=3, value=f["base"])
            c_m = ws.cell(row=row_num, column=4, value=f["monto"])

            if f["unidad"] is not None:
                c_u.number_format = '0.00%' if abs(f["unidad"]) < 1 else '0.00'
            c_b.number_format = '$#,##0.00'
            c_m.number_format = '$#,##0.00'

            c_desc.font = font_bold if es_total else font_normal
            c_u.font = font_bold if es_total else font_normal
            c_b.font = font_bold if es_total else font_normal
            c_m.font = font_bold if es_total else font_normal

            c_desc.alignment = align_left
            c_u.alignment = align_right
            c_b.alignment = align_right
            c_m.alignment = align_right

            row_num += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    if path_grafico and os.path.exists(path_grafico):
        img = OpenpyxlImage(path_grafico)
        img.width = 450
        img.height = 338
        ws.add_image(img, "F2")

    wb.save(path)


# ======================================================================
# Exportar Recibo: ODS (odfpy)
# ======================================================================
def exportar_recibo_ods(resultado: dict, db, path: str, path_grafico: str | None):
    """Exporta el recibo detallado a un documento ODS."""
    doc = OpenDocumentSpreadsheet()
    table = Table(name="Recibo de Sueldo")

    emp = resultado["empleado"]

    def add_row_info(label, val):
        row = TableRow()
        c1 = TableCell()
        c1.addElement(P(text=label))
        row.addElement(c1)
        c2 = TableCell()
        c2.addElement(P(text=val))
        row.addElement(c2)
        table.addElement(row)

    r_tit = TableRow()
    c_tit = TableCell()
    c_tit.addElement(P(text="RECIBO DE SUELDO Y COSTO LABORAL"))
    r_tit.addElement(c_tit)
    table.addElement(r_tit)

    table.addElement(TableRow())

    add_row_info("Legajo:", emp["legajo"])
    add_row_info("Empleado:", emp["nombre_completo"])
    add_row_info("Tipo Liquidación:", emp["tipo_liquidacion"].capitalize())
    if resultado.get("quincena_sel"):
        add_row_info("Quincena:", resultado["quincena_sel"])

    table.addElement(TableRow())

    r_h = TableRow()
    for h in ["Concepto", "Unidad / %", "Base", "Monto"]:
        cell = TableCell()
        cell.addElement(P(text=h))
        r_h.addElement(cell)
    table.addElement(r_h)

    secciones_info = {s["codigo"]: s["titulo"] for s in db.listar_secciones()}
    orden_secciones = [s["codigo"] for s in db.listar_secciones()]
    for s_cod in resultado["resultados_por_seccion"].keys():
        if s_cod not in orden_secciones:
            orden_secciones.append(s_cod)

    for sec_codigo in orden_secciones:
        filas = resultado["resultados_por_seccion"].get(sec_codigo, [])
        filas_visibles = []
        for f in filas:
            if f.get("visible_recibo", 1) == 1:
                es_total = f["codigo"].startswith("total_") or f["codigo"] in ("bruto", "total_deducciones", "neto", "total_cargas_patronales", "costo_laboral_total")
                if f["monto"] != 0 or es_total:
                    filas_visibles.append(f)
        if not filas_visibles:
            continue

        r_sec = TableRow()
        c_sec = TableCell()
        c_sec.addElement(P(text=secciones_info.get(sec_codigo, sec_codigo).upper()))
        r_sec.addElement(c_sec)
        table.addElement(r_sec)

        for f in filas_visibles:
            row = TableRow()

            c_desc = TableCell()
            c_desc.addElement(P(text=f["descripcion"]))
            row.addElement(c_desc)

            c_u = TableCell()
            c_u.addElement(P(text=_formato_porcentaje(f["unidad"]) if f["unidad"] is not None else ""))
            row.addElement(c_u)

            c_b = TableCell()
            c_b.addElement(P(text=_formato_moneda(f["base"]) if f["base"] is not None else ""))
            row.addElement(c_b)

            c_m = TableCell()
            c_m.addElement(P(text=_formato_moneda(f["monto"])))
            row.addElement(c_m)

            table.addElement(row)

    if path_grafico and os.path.exists(path_grafico):
        href = doc.addPicture(path_grafico)
        r_img = TableRow()
        c_img = TableCell()
        frame = Frame(width="12cm", height="9cm", x="0cm", y="0cm")
        frame.addElement(Image(href=href))
        c_img.addElement(frame)
        r_img.addElement(c_img)
        table.addElement(r_img)

    doc.spreadsheet.addElement(table)
    doc.save(path)