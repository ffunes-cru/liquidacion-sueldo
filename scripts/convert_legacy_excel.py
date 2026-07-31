#!/usr/bin/env python3
import os
import sys
import shutil
import openpyxl

def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_dir, "datos_liquidacion_sueldos.xlsx")
    backup_path = os.path.join(base_dir, "datos_liquidacion_sueldos.bak.xlsx")
    output_path = os.path.join(base_dir, "datos_liquidacion_sueldos_compatible.xlsx")

    if len(sys.argv) > 1:
        excel_path = os.path.abspath(sys.argv[1])

    if not os.path.exists(excel_path):
        print(f"Error: No se encontró el archivo Excel en {excel_path}")
        sys.exit(1)

    print(f"Cargando archivo Excel heredado: {excel_path}")
    if os.path.abspath(excel_path) != os.path.abspath(backup_path):
        shutil.copyfile(excel_path, backup_path)
        print(f"Respaldo creado en: {backup_path}")

    wb = openpyxl.load_workbook(excel_path)
    
    # ── 1. Esquemas de Cálculo ─────────────────────────────────────
    schemas = []
    if "Esquemas de Cálculo" in wb.sheetnames:
        ws = wb["Esquemas de Cálculo"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c).lower().strip() if c else "" for c in rows[0]]
            code_idx = header.index("codigo") if "codigo" in header else 0
            name_idx = header.index("nombre") if "nombre" in header else 1
            tipo_idx = header.index("tipo_liquidacion") if "tipo_liquidacion" in header else -1
            
            for r in rows[1:]:
                if r and r[code_idx]:
                    code = str(r[code_idx]).strip().upper()
                    name = str(r[name_idx]).strip() if len(r) > name_idx and r[name_idx] else code
                    tipo = str(r[tipo_idx]).strip().lower() if tipo_idx >= 0 and len(r) > tipo_idx and r[tipo_idx] else ("jornal" if code == "JORNAL" else "mensual")
                    schemas.append({"codigo": code, "nombre": name, "tipo_liquidacion": tipo})

    if not schemas:
        schemas = [
            {"codigo": "MENSUAL", "nombre": "Comercio Mensualizado", "tipo_liquidacion": "mensual"},
            {"codigo": "JORNAL", "nombre": "Comercio Jornalero (Por hora)", "tipo_liquidacion": "jornal"}
        ]

    # ── 2. Categorías Jornaleras ──────────────────────────────────
    categories = []
    if "Categorías Jornaleras" in wb.sheetnames:
        ws = wb["Categorías Jornaleras"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            for r in rows[1:]:
                if r and r[0] is not None:
                    try:
                        cat_id = int(r[0])
                        name = str(r[1]).strip() if len(r) > 1 and r[1] else f"Categoría {cat_id}"
                        valor = float(r[2]) if len(r) > 2 and r[2] is not None else 0.0
                        categories.append({"id": cat_id, "nombre": name, "valor_hora": valor})
                    except (ValueError, TypeError):
                        pass

    # ── 3. Secciones ─────────────────────────────────────────────
    sections = []
    if "Secciones" in wb.sheetnames:
        ws = wb["Secciones"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            for r in rows[1:]:
                if r and r[0] is not None:
                    try:
                        sec_id = int(r[0])
                        code = str(r[1]).strip() if len(r) > 1 and r[1] else ""
                        title = str(r[2]).strip() if len(r) > 2 and r[2] else code
                        orden = int(r[3]) if len(r) > 3 and r[3] is not None else (sec_id * 10)
                        sections.append({"id": sec_id, "codigo": code, "titulo": title, "orden": orden})
                    except (ValueError, TypeError):
                        pass

    # ── 4. Celdas de Cálculo ─────────────────────────────────────
    cells = []
    if "Celdas de Cálculo" in wb.sheetnames:
        ws = wb["Celdas de Cálculo"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            for r in rows[1:]:
                if r and r[0] is not None:
                    try:
                        cell_id = int(r[0])
                        sec_code = str(r[1]).strip() if len(r) > 1 and r[1] else "COMPOSICION"
                        var_code = str(r[2]).strip().lower() if len(r) > 2 and r[2] else ""
                        desc = str(r[3]).strip() if len(r) > 3 and r[3] else ""
                        cond = str(r[4]).strip() if len(r) > 4 and r[4] else ""
                        f_unid = str(r[5]).strip() if len(r) > 5 and r[5] else ""
                        f_base = str(r[6]).strip() if len(r) > 6 and r[6] else ""
                        f_monto = str(r[7]).strip() if len(r) > 7 and r[7] else ""
                        orden = int(r[8]) if len(r) > 8 and r[8] is not None else 0
                        esq_code = str(r[9]).strip().upper() if len(r) > 9 and r[9] else "MENSUAL"
                        t_calc = str(r[10]).strip().lower() if len(r) > 10 and r[10] else "formula"
                        s_porc = float(r[11]) if len(r) > 11 and r[11] is not None else 0.0
                        s_base = str(r[12]).strip().lower() if len(r) > 12 and r[12] else ""
                        s_monto = float(r[13]) if len(r) > 13 and r[13] is not None else 0.0
                        cells.append({
                            "id": cell_id, "seccion_codigo": sec_code, "codigo_variable": var_code,
                            "descripcion": desc, "condicion": cond, "formula_unidad": f_unid,
                            "formula_base": f_base, "formula_monto": f_monto, "orden": orden,
                            "esquema_codigo": esq_code, "tipo_calculo": t_calc, "simple_porcentaje": s_porc,
                            "simple_base_variable": s_base, "simple_monto_fijo": s_monto
                        })
                    except (ValueError, TypeError):
                        pass

    # ── 5. Process Empleados & Custom Fields ──────────────────────
    employees = []
    schema_fields = []
    quincenas_emp = []
    field_values = []

    # Read existing relational sheets if already populated
    if "Variables de Esquema" in wb.sheetnames:
        ws = wb["Variables de Esquema"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            for r in rows[1:]:
                if r and r[0] is not None:
                    try:
                        sf_id = int(r[0])
                        esq = str(r[1]).strip().upper() if len(r) > 1 and r[1] else "JORNAL"
                        code = str(r[2]).strip() if len(r) > 2 and r[2] else ""
                        label = str(r[3]).strip() if len(r) > 3 and r[3] else code.replace("_", " ").title()
                        ftype = str(r[4]).strip() if len(r) > 4 and r[4] else "number"
                        defval = str(r[5]).strip() if len(r) > 5 and r[5] is not None else "0"
                        disp_ord = int(r[6]) if len(r) > 6 and r[6] is not None else sf_id * 10
                        if code:
                            schema_fields.append({
                                "id": sf_id, "esquema_codigo": esq, "field_code": code,
                                "field_label": label, "field_type": ftype,
                                "default_value": defval, "display_order": disp_ord
                            })
                    except (ValueError, TypeError):
                        pass

    if "Quincenas Empleado" in wb.sheetnames:
        ws = wb["Quincenas Empleado"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            for r in rows[1:]:
                if r and r[0] is not None:
                    try:
                        emp_id = int(r[0])
                        qn = str(r[1]).strip() if len(r) > 1 and r[1] else "Q1"
                        quincenas_emp.append({"empleado_id": emp_id, "quincena": qn})
                    except (ValueError, TypeError):
                        pass

    if "Valores de Empleados" in wb.sheetnames:
        ws = wb["Valores de Empleados"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            for r in rows[1:]:
                if r and r[0] is not None:
                    try:
                        fv_id = int(r[0])
                        emp_id = int(r[1])
                        f_id = int(r[2])
                        qn = str(r[3]).strip() if len(r) > 3 and r[3] else "Q1"
                        val = str(r[4]).strip() if len(r) > 4 and r[4] is not None else "0"
                        field_values.append({
                            "id": fv_id, "empleado_id": emp_id, "field_id": f_id,
                            "quincena": qn, "value": val
                        })
                    except (ValueError, TypeError):
                        pass

    field_map = {} # field_code -> field_id
    field_counter = len(schema_fields) + 1

    # Ensure MENSUAL schema has input fields
    mensual_default_fields = [
        ("sueldo", "Sueldo Básico", "number", "0"),
        ("horas_extras_50", "Horas Extras 50%", "number", "0"),
        ("horas_extras_100", "Horas Extras 100%", "number", "0"),
        ("ret_ingresos_rel_rep", "Ret. Ing. Rel. Dep. Act.", "number", "0"),
        ("ajret_ingresos_rel_dep", "Aj. Ret. Ing. Rel. Dep. Ant.", "number", "0"),
        ("devret_ingresos_rel_dep", "Dev. Ret. Ing. Rel. Dep. Ant.", "number", "0"),
    ]

    for code, label, ftype, defval in mensual_default_fields:
        if not any(sf["esquema_codigo"] == "MENSUAL" and sf["field_code"] == code for sf in schema_fields):
            sf_id = field_counter
            field_counter += 1
            schema_fields.append({
                "id": sf_id,
                "esquema_codigo": "MENSUAL",
                "field_code": code,
                "field_label": label,
                "field_type": ftype,
                "default_value": defval,
                "display_order": sf_id * 10
            })

    # Build initial field_map for all schema_fields
    for sf in schema_fields:
        field_map[sf["field_code"]] = sf["id"]

    if "Empleados" in wb.sheetnames:
        ws = wb["Empleados"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(c).strip() if c is not None else "" for c in rows[0]]
            base_col_count = 8 # id, legajo, nombre_completo, tipo_liquidacion, esquema_codigo, categoria_jornal_id, fecha_ingreso, cuil
            
            dynamic_cols = []
            for col_idx in range(base_col_count, len(headers)):
                h = headers[col_idx]
                if h:
                    clean_code = h[2:] if h.startswith("j_") or h.startswith("m_") else h
                    dynamic_cols.append((col_idx, clean_code, h))
                    
                    if clean_code not in field_map:
                        field_id = field_counter
                        field_map[clean_code] = field_id
                        field_counter += 1
                        
                        label = clean_code.replace("_", " ").title()

                        # Infer field_type from cell values
                        inferred_type = "number"
                        col_vals = [r[col_idx] for r in rows[1:] if len(r) > col_idx and r[col_idx] is not None]
                        if col_vals:
                            bool_count = 0
                            num_count = 0
                            str_count = 0
                            for v in col_vals:
                                if isinstance(v, bool) or str(v).strip().lower() in ["true", "false", "si", "no"]:
                                    bool_count += 1
                                else:
                                    try:
                                        float(str(v).replace(',', '.'))
                                        num_count += 1
                                    except ValueError:
                                        str_count += 1
                            if bool_count > 0 and num_count == 0 and str_count == 0:
                                inferred_type = "bool"
                            elif str_count > num_count:
                                inferred_type = "string"
                            else:
                                inferred_type = "number"

                        schema_fields.append({
                            "id": field_id,
                            "esquema_codigo": "JORNAL" if h.startswith("j_") else "MENSUAL",
                            "field_code": clean_code,
                            "field_label": label,
                            "field_type": inferred_type,
                            "default_value": "false" if inferred_type == "bool" else ("0" if inferred_type == "number" else ""),
                            "display_order": field_id * 10
                        })

            current_emp_id = 0
            val_id_counter = len(field_values) + 1

            for r in rows[1:]:
                if not r or r[0] is None:
                    continue
                raw_id = str(r[0]).strip()
                if raw_id.isdigit() and int(raw_id) > 0:
                    current_emp_id = int(raw_id)
                    legajo = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
                    nombre = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
                    tipo_liq = str(r[3]).strip().lower() if len(r) > 3 and r[3] is not None else "mensual"
                    esquema = str(r[4]).strip().upper() if len(r) > 4 and r[4] is not None else "MENSUAL"
                    cat_id = int(r[5]) if len(r) > 5 and r[5] is not None and str(r[5]).isdigit() else 0
                    fecha_ing = str(r[6]).strip() if len(r) > 6 and r[6] is not None else "2020-01-01"
                    cuil = str(r[7]).strip() if len(r) > 7 and r[7] is not None else ""

                    employees.append({
                        "id": current_emp_id,
                        "legajo": legajo,
                        "nombre_completo": nombre,
                        "tipo_liquidacion": tipo_liq,
                        "esquema_codigo": esquema,
                        "categoria_jornal_id": cat_id,
                        "fecha_ingreso": fecha_ing,
                        "cuil": cuil
                    })

                    if not any(qe["empleado_id"] == current_emp_id for qe in quincenas_emp):
                        quincena_code = "Q1"
                        quincenas_emp.append({"empleado_id": current_emp_id, "quincena": quincena_code})

                    for col_idx, clean_code, raw_h in dynamic_cols:
                        if col_idx < len(r) and r[col_idx] is not None:
                            val = str(r[col_idx]).strip()
                            if val != "" and val != "0":
                                field_values.append({
                                    "id": val_id_counter,
                                    "empleado_id": current_emp_id,
                                    "field_id": field_map[clean_code],
                                    "quincena": "Q1",
                                    "value": val
                                })
                                val_id_counter += 1
                elif current_emp_id > 0 and dynamic_cols:
                    # Sub-row representing Q2
                    quincena_code = "Q2"
                    if not any(qe["empleado_id"] == current_emp_id and qe["quincena"] == "Q2" for qe in quincenas_emp):
                        quincenas_emp.append({"empleado_id": current_emp_id, "quincena": quincena_code})

                    for col_idx, clean_code, raw_h in dynamic_cols:
                        if col_idx < len(r) and r[col_idx] is not None:
                            val = str(r[col_idx]).strip()
                            if val != "" and val != "0":
                                field_values.append({
                                    "id": val_id_counter,
                                    "empleado_id": current_emp_id,
                                    "field_id": field_map[clean_code],
                                    "quincena": quincena_code,
                                    "value": val
                                })
                                val_id_counter += 1

    # ── 6. Variables Globales, Empresa, Recibos ────────────────────
    globals_list = []
    if "Variables Globales" in wb.sheetnames:
        ws = wb["Variables Globales"]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            if r and r[0] is not None:
                globals_list.append({
                    "id": int(r[0]), "codigo": str(r[1]).strip(),
                    "valor": str(r[2]).strip(), "descripcion": str(r[3]).strip() if len(r) > 3 and r[3] else ""
                })

    company = {}
    if "Empresa" in wb.sheetnames:
        ws = wb["Empresa"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1 and rows[1] and rows[1][0] is not None:
            r = rows[1]
            company = {
                "id": int(r[0]), "razon_social": str(r[1]).strip() if len(r) > 1 and r[1] else "",
                "direccion": str(r[2]).strip() if len(r) > 2 and r[2] else "",
                "cuit": str(r[3]).strip() if len(r) > 3 and r[3] else "",
                "lugar_de_pago": str(r[4]).strip() if len(r) > 4 and r[4] else ""
            }

    recibos_list = []
    if "Recibos" in wb.sheetnames:
        ws = wb["Recibos"]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            if r and r[0] is not None:
                try:
                    recibos_list.append({
                        "id": int(r[0]), "empleado_id": int(r[1]), "esquema_codigo": str(r[2]).strip(),
                        "mes": int(r[3]), "anio": int(r[4]), "periodo": str(r[5]).strip(),
                        "datos_json": str(r[6]).strip(), "fecha_emision": str(r[7]).strip() if len(r) > 7 and r[7] else ""
                    })
                except (ValueError, TypeError):
                    pass

    # ── BUILD NEW COMPATIBLE WORKBOOK ──────────────────────────────
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active) # Remove default sheet

    # 1. Esquemas de Cálculo
    ws1 = out_wb.create_sheet("Esquemas de Cálculo")
    ws1.append(["codigo", "nombre", "tipo_liquidacion"])
    for s in schemas:
        ws1.append([s["codigo"], s["nombre"], s["tipo_liquidacion"]])

    # 2. Categorías Jornaleras
    ws2 = out_wb.create_sheet("Categorías Jornaleras")
    ws2.append(["id", "nombre", "valor_hora"])
    for c in categories:
        ws2.append([c["id"], c["nombre"], c["valor_hora"]])

    # 3. Secciones
    ws3 = out_wb.create_sheet("Secciones")
    ws3.append(["id", "codigo", "titulo", "orden"])
    for sec in sections:
        ws3.append([sec["id"], sec["codigo"], sec["titulo"], sec["orden"]])

    # 4. Empleados
    ws4 = out_wb.create_sheet("Empleados")
    ws4.append(["id", "legajo", "nombre_completo", "tipo_liquidacion", "esquema_codigo", "categoria_jornal_id", "fecha_ingreso", "cuil"])
    for e in employees:
        ws4.append([e["id"], e["legajo"], e["nombre_completo"], e["tipo_liquidacion"], e["esquema_codigo"], e["categoria_jornal_id"], e["fecha_ingreso"], e["cuil"]])

    # 5. Celdas de Cálculo
    ws5 = out_wb.create_sheet("Celdas de Cálculo")
    ws5.append(["id", "seccion_codigo", "codigo_variable", "descripcion", "condicion", "formula_unidad", "formula_base", "formula_monto", "orden", "esquema_codigo", "tipo_calculo", "simple_porcentaje", "simple_base_variable", "simple_monto_fijo"])
    for cell in cells:
        ws5.append([
            cell["id"], cell["seccion_codigo"], cell["codigo_variable"], cell["descripcion"],
            cell["condicion"], cell["formula_unidad"], cell["formula_base"], cell["formula_monto"],
            cell["orden"], cell["esquema_codigo"], cell["tipo_calculo"], cell["simple_porcentaje"],
            cell["simple_base_variable"], cell["simple_monto_fijo"]
        ])

    # 6. Variables Globales
    ws6 = out_wb.create_sheet("Variables Globales")
    ws6.append(["id", "codigo", "valor", "descripcion"])
    for g in globals_list:
        ws6.append([g["id"], g["codigo"], g["valor"], g["descripcion"]])

    # 7. Empresa
    ws7 = out_wb.create_sheet("Empresa")
    ws7.append(["id", "razon_social", "direccion", "cuit", "lugar_de_pago"])
    if company:
        ws7.append([company["id"], company["razon_social"], company["direccion"], company["cuit"], company["lugar_de_pago"]])

    # 8. Variables de Esquema
    ws8 = out_wb.create_sheet("Variables de Esquema")
    ws8.append(["id", "esquema_codigo", "field_code", "field_label", "field_type", "default_value", "display_order"])
    for sf in schema_fields:
        ws8.append([sf["id"], sf["esquema_codigo"], sf["field_code"], sf["field_label"], sf["field_type"], sf["default_value"], sf["display_order"]])

    # 9. Quincenas Empleado
    ws9 = out_wb.create_sheet("Quincenas Empleado")
    ws9.append(["empleado_id", "quincena"])
    for qe in quincenas_emp:
        ws9.append([qe["empleado_id"], qe["quincena"]])

    # 10. Valores de Empleados
    ws10 = out_wb.create_sheet("Valores de Empleados")
    ws10.append(["id", "empleado_id", "field_id", "quincena", "value"])
    for ev in field_values:
        ws10.append([ev["id"], ev["empleado_id"], ev["field_id"], ev["quincena"], ev["value"]])

    # Save to both target output path and original excel path
    out_wb.save(output_path)
    out_wb.save(excel_path)

    print("==========================================================================")
    print(" CONVERSIÓN COMPLETADA EXITOSAMENTE")
    print(f" - Empleados procesados: {len(employees)}")
    print(f" - Variables de Esquema creadas: {len(schema_fields)}")
    print(f" - Asignaciones Quincenales: {len(quincenas_emp)}")
    print(f" - Valores de Campos registrados: {len(field_values)}")
    print(f" - Archivo Excel actualizado: {excel_path}")
    print(f" - Copia compatible guardada en: {output_path}")
    print("==========================================================================")

if __name__ == "__main__":
    main()
